# -*- coding: utf-8 -*-
from __future__ import annotations

import copy
import importlib.util
import json
import os
import re
import shutil
import subprocess
import sys
import unittest
import zipfile
from pathlib import Path

ROOT = Path(__file__).resolve().parents[2]
SKILL_DIR = ROOT / "skills" / "prospectus-extractor"
SCRIPTS = SKILL_DIR / "scripts"
sys.path.insert(0, str(SCRIPTS))

from build_manifest import build_manifest, inspect_xlsx
from build_excel import build
from factbase_from_case import render_text as render_factbase
from prospectus_contract import COVERAGE_ITEMS, SHEET_SKELETON_NAMES
from render_case_markdown import render_coverage, render_raw
from slice_prospectus import assign_end_pages
from validate_case_data import validate_data


def write_blank_pdf(path: Path) -> None:
    from pypdf import PdfWriter

    writer = PdfWriter()
    writer.add_blank_page(width=612, height=792)
    with path.open("wb") as handle:
        writer.write(handle)


class ContractTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls) -> None:
        cls.fixture_path = SCRIPTS / "example_case_data.json"
        cls.data = json.loads(cls.fixture_path.read_text(encoding="utf-8"))

    def test_schema_document_is_valid_json(self) -> None:
        schema = json.loads((SKILL_DIR / "references" / "case_data.schema.json").read_text(encoding="utf-8"))
        self.assertEqual(schema["properties"]["schema_version"]["const"], "1.0")
        self.assertEqual(schema["properties"]["coverage"]["minItems"], 24)
        self.assertEqual(schema["properties"]["coverage"]["maxItems"], 24)

    def test_skill_and_references_use_current_contract(self) -> None:
        skill = (SKILL_DIR / "SKILL.md").read_text(encoding="utf-8")
        integration = (SKILL_DIR / "references" / "integration.md").read_text(encoding="utf-8")
        excel = (SKILL_DIR / "references" / "excel_output.md").read_text(encoding="utf-8")
        combined = "\n".join((skill, integration, excel))
        self.assertNotIn("/ 16", combined)
        self.assertNotIn(".claude", combined)
        self.assertIn("5 個業務產物＋1 個控制檔", skill)
        self.assertIn("case_data.schema.json", skill)
        self.assertIn("29_AI專家", excel)
        self.assertIn("紅旗", excel)
        for reference in (
            "references/section_map.md",
            "references/case_data.schema.json",
            "references/excel_output.md",
            "references/integration.md",
        ):
            self.assertTrue((SKILL_DIR / reference).is_file(), reference)

    def test_example_passes_contract(self) -> None:
        result = validate_data(self.data)
        self.assertEqual(result["status"], "success", result["errors"])
        self.assertEqual(result["coverage_items"], 24)
        self.assertEqual(result["skeleton_sheets"], 35)

    def test_section_map_has_canonical_w_ids_and_titles(self) -> None:
        lines = (SKILL_DIR / "references" / "section_map.md").read_text(encoding="utf-8").splitlines()
        mapped = []
        for line in lines:
            if re.match(r"^\| W\d{2} \|", line):
                cells = [cell.strip() for cell in line.strip("|").split("|")]
                mapped.append((cells[0], cells[1]))
        self.assertEqual(mapped, COVERAGE_ITEMS)

    def test_validator_normalizes_fullwidth_and_ascii_slashes(self) -> None:
        data = copy.deepcopy(self.data)
        data["coverage"][16]["title"] = "主要銷貨/客戶"
        data["coverage"][19]["title"] = "產品別/部門別毛利率變化"
        result = validate_data(data)
        self.assertEqual(result["status"], "success", result["errors"])

    def test_missing_coverage_item_fails(self) -> None:
        data = copy.deepcopy(self.data)
        data["coverage"].pop()
        result = validate_data(data)
        self.assertEqual(result["status"], "failed")
        self.assertTrue(any("exactly 24" in error for error in result["errors"]))

    def test_reordered_core_sheet_fails(self) -> None:
        data = copy.deepcopy(self.data)
        data["sheets"][2], data["sheets"][3] = data["sheets"][3], data["sheets"][2]
        result = validate_data(data)
        self.assertEqual(result["status"], "failed")
        self.assertTrue(any("35 canonical sheets in order" in error for error in result["errors"]))

    def test_not_applicable_is_not_missing(self) -> None:
        data = copy.deepcopy(self.data)
        item = data["coverage"][17]
        self.assertEqual(item["status"], "not_applicable")
        coverage = render_coverage(data)
        self.assertIn("not_applicable 1", coverage)
        self.assertIn("⚪ 不適用", coverage)

    def test_markdown_outputs_derive_from_case_data(self) -> None:
        raw = render_raw(self.data)
        coverage = render_coverage(self.data)
        self.assertIn("SRC-001 / physical p.1", raw)
        self.assertIn("total 24", coverage)
        self.assertEqual(coverage.count("\n| W"), 24)


class SlicerTests(unittest.TestCase):
    runtime_dir = Path(__file__).parent / "_runtime_pdf"

    def tearDown(self) -> None:
        shutil.rmtree(self.runtime_dir, ignore_errors=True)

    def test_nested_child_does_not_truncate_parent(self) -> None:
        items = [
            (0, "壹、公司概況", 5),
            (1, "一、公司簡介", 6),
            (2, "（一）設立日期", 7),
            (1, "二、風險事項", 12),
            (0, "貳、營運概況", 20),
        ]
        sections = assign_end_pages(items, 30)
        by_title = {section["title"]: section for section in sections}
        self.assertEqual(by_title["壹、公司概況"]["end_page"], 19)
        self.assertEqual(by_title["一、公司簡介"]["end_page"], 11)
        self.assertEqual(by_title["（一）設立日期"]["end_page"], 11)
        self.assertEqual(by_title["二、風險事項"]["end_page"], 19)
        self.assertEqual(by_title["貳、營運概況"]["end_page"], 30)

    def test_same_page_child_keeps_valid_range(self) -> None:
        items = [(0, "壹、公司概況", 5), (1, "一、公司簡介", 5), (0, "貳、營運概況", 10)]
        sections = assign_end_pages(items, 12)
        self.assertEqual(sections[0]["start_page"], 5)
        self.assertEqual(sections[0]["end_page"], 9)
        self.assertEqual(sections[1]["end_page"], 9)

    def test_step1_works_without_poppler_and_renderer_is_explicit(self) -> None:
        self.runtime_dir.mkdir(exist_ok=True)
        pdf = self.runtime_dir / "one-page.pdf"
        index_dir = self.runtime_dir / "index"
        image_dir = self.runtime_dir / "images"
        write_blank_pdf(pdf)
        env = os.environ.copy()
        env["PATH"] = ""
        env["PYTHONDONTWRITEBYTECODE"] = "1"

        sliced = subprocess.run(
            [
                sys.executable,
                "-X",
                "utf8",
                str(SCRIPTS / "slice_prospectus.py"),
                str(pdf),
                "--outdir",
                str(index_dir),
                "--dump-sections",
            ],
            capture_output=True,
            text=True,
            encoding="utf-8",
            errors="replace",
            env=env,
            check=False,
        )
        self.assertEqual(sliced.returncode, 0, sliced.stderr)
        index = json.loads((index_dir / "index.json").read_text(encoding="utf-8"))
        self.assertEqual(index["meta"]["pages"], 1)
        self.assertEqual(index["meta"]["scanned_page_count"], 1)

        rendered = subprocess.run(
            [
                sys.executable,
                "-X",
                "utf8",
                str(SCRIPTS / "render_pdf_pages.py"),
                str(pdf),
                "--outdir",
                str(image_dir),
                "--pages",
                "1",
            ],
            capture_output=True,
            text=True,
            encoding="utf-8",
            errors="replace",
            env=env,
            check=False,
        )
        if importlib.util.find_spec("fitz"):
            self.assertEqual(rendered.returncode, 0, rendered.stderr)
            payload = json.loads(rendered.stdout)
            self.assertEqual(payload["backend"], "pymupdf")
            self.assertTrue((image_dir / "page-0001.png").is_file())
        else:
            self.assertEqual(rendered.returncode, 2)
            self.assertIn("缺少 PDF 轉圖後端", rendered.stderr)
            self.assertIn("pip install -r requirements.txt", rendered.stderr)
            self.assertNotIn("Traceback", rendered.stderr)

    def test_renderer_uses_available_backend_on_normal_path(self) -> None:
        self.runtime_dir.mkdir(exist_ok=True)
        pdf = self.runtime_dir / "one-page.pdf"
        image_dir = self.runtime_dir / "normal-path-images"
        write_blank_pdf(pdf)
        env = os.environ.copy()
        env["PYTHONDONTWRITEBYTECODE"] = "1"
        rendered = subprocess.run(
            [
                sys.executable,
                "-X",
                "utf8",
                str(SCRIPTS / "render_pdf_pages.py"),
                str(pdf),
                "--outdir",
                str(image_dir),
                "--pages",
                "1",
            ],
            capture_output=True,
            text=True,
            encoding="utf-8",
            errors="replace",
            env=env,
            check=False,
        )
        self.assertEqual(rendered.returncode, 0, rendered.stderr)
        payload = json.loads(rendered.stdout)
        self.assertIn(payload["backend"], ("pymupdf", "pdftoppm"))
        self.assertTrue((image_dir / "page-0001.png").is_file())

    def test_step1_missing_pypdf_has_readable_install_message(self) -> None:
        self.runtime_dir.mkdir(exist_ok=True)
        pdf = self.runtime_dir / "one-page.pdf"
        write_blank_pdf(pdf)
        env = os.environ.copy()
        env["PATH"] = ""
        env["PYTHONDONTWRITEBYTECODE"] = "1"
        result = subprocess.run(
            [
                sys.executable,
                "-S",
                str(SCRIPTS / "slice_prospectus.py"),
                str(pdf),
                "--outdir",
                str(self.runtime_dir / "index"),
            ],
            capture_output=True,
            text=True,
            encoding="utf-8",
            errors="replace",
            env=env,
            check=False,
        )
        self.assertEqual(result.returncode, 2)
        self.assertIn("缺少必要 Python 套件 pypdf", result.stderr)
        self.assertIn("pip install -r requirements.txt", result.stderr)
        self.assertNotIn("Traceback", result.stderr)


class ManifestExcelInspectionTests(unittest.TestCase):
    runtime_dir = Path(__file__).parent / "_runtime_manifest"

    def tearDown(self) -> None:
        shutil.rmtree(self.runtime_dir, ignore_errors=True)

    def test_ooxml_sheet_order_without_openpyxl(self) -> None:
        namespace = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
        sheets = "".join(
            f'<sheet name="{name}" sheetId="{index + 1}" r:id="rId{index + 1}"/>'
            for index, name in enumerate(SHEET_SKELETON_NAMES)
        )
        workbook_xml = (
            '<?xml version="1.0" encoding="UTF-8"?>'
            f'<workbook xmlns="{namespace}" '
            'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">'
            f"<sheets>{sheets}</sheets></workbook>"
        )
        path = Path(__file__).parent / "_manifest_fixture.xlsx"
        try:
            with zipfile.ZipFile(path, "w") as archive:
                archive.writestr("xl/workbook.xml", workbook_xml)
            result = inspect_xlsx(path)
        finally:
            path.unlink(missing_ok=True)
        self.assertEqual(result["status"], "success", result)
        self.assertEqual(result["sheets"], 35)

    def test_manifest_truthfully_fails_for_synthetic_fixture(self) -> None:
        manifest = build_manifest(SCRIPTS / "example_case_data.json", SCRIPTS)
        self.assertEqual(manifest["validation_status"], "failed")
        self.assertFalse(manifest["qa"]["inputs_available"])
        self.assertEqual(manifest["manifest_hash_policy"], "self_excluded_to_avoid_recursive_hash")

    def test_manifest_uses_pypdf_and_rejects_garbage_markdown(self) -> None:
        self.runtime_dir.mkdir(exist_ok=True)
        data = json.loads((SCRIPTS / "example_case_data.json").read_text(encoding="utf-8"))
        case_id = "manifest_case"
        pdf = self.runtime_dir / "prospectus.pdf"
        write_blank_pdf(pdf)
        data["case"]["case_id"] = case_id
        data["sources"][0]["path_or_url"] = str(pdf)
        data["sources"][0]["snapshot_path"] = str(pdf)
        case_data = self.runtime_dir / f"{case_id}_case_data.json"
        case_data.write_text(json.dumps(data, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
        (self.runtime_dir / f"{case_id}_Prospectus_raw.md").write_bytes(render_raw(data).encode("utf-8"))
        (self.runtime_dir / f"{case_id}_Factbase.md").write_bytes(render_factbase(data).encode("utf-8"))
        coverage_path = self.runtime_dir / f"{case_id}_Prospectus_coverage.md"
        coverage_path.write_bytes(render_coverage(data).encode("utf-8"))
        build(data, self.runtime_dir / f"{case_id}_Prospectus_extract.xlsx")

        good = build_manifest(case_data, self.runtime_dir)
        self.assertEqual(good["validation_status"], "success", good)
        self.assertEqual(good["inputs"][0]["pages"], 1)
        self.assertEqual(good["inputs"][0]["page_count_backend"], "pypdf")
        self.assertTrue(good["qa"]["excel"]["coverage_consistent"])

        (self.runtime_dir / f"{case_id}_Prospectus_raw.md").write_bytes(
            "Raw 也被替換成垃圾\n".encode("utf-8")
        )
        coverage_path.write_text("這是完全無關的垃圾內容\n", encoding="utf-8")
        (self.runtime_dir / f"{case_id}_Factbase.md").write_text(
            "同樣把 Factbase 換成垃圾\n", encoding="utf-8"
        )
        bad = build_manifest(case_data, self.runtime_dir)
        self.assertEqual(bad["validation_status"], "failed")
        by_name = {Path(item["path"]).name: item for item in bad["outputs"]}
        self.assertFalse(by_name[f"{case_id}_Prospectus_raw.md"]["content_match"])
        self.assertFalse(by_name[f"{case_id}_Prospectus_coverage.md"]["content_match"])
        self.assertFalse(by_name[f"{case_id}_Factbase.md"]["content_match"])
        self.assertFalse(bad["qa"]["rendered_outputs_match"])


class LegacyExcelCompatibilityTests(unittest.TestCase):
    def test_builder_distinguishes_missing_and_not_applicable(self) -> None:
        data = json.loads((SCRIPTS / "example_case_data.json").read_text(encoding="utf-8"))
        path = Path(__file__).parent / "_legacy_fixture.xlsx"
        try:
            names = build(data, path)
            self.assertEqual(names, SHEET_SKELETON_NAMES)
            from openpyxl import load_workbook

            workbook = load_workbook(path, read_only=True, data_only=False)
            self.assertIn("查無資料／須補件", workbook["02_股本形成"]["A2"].value)
            self.assertIn("不適用", workbook["29_AI專家"]["A2"].value)
            self.assertEqual(workbook["00_覆蓋率"].max_row, 25)
            self.assertEqual(workbook["00_覆蓋率"]["A2"].value, "W01")
            self.assertEqual(workbook["00_覆蓋率"]["A25"].value, "W24")
            workbook.close()
            verified = subprocess.run(
                [sys.executable, str(SCRIPTS / "verify_excel.py"), str(path)],
                capture_output=True,
                text=True,
                encoding="utf-8",
                errors="replace",
                check=False,
            )
            self.assertEqual(verified.returncode, 0, verified.stderr)
            payload = json.loads(verified.stdout)
            self.assertNotIn("00_覆蓋率", payload["flagged_not_applicable_pages"])
            self.assertEqual(
                payload["flagged_not_applicable_pages"],
                ["19_銷售量值", "20_生產量值", "21_產能利用率", "29_AI專家"],
            )
        finally:
            path.unlink(missing_ok=True)

    def test_eq_verifier_no_longer_uses_eval(self) -> None:
        source = (SCRIPTS / "verify_excel.py").read_text(encoding="utf-8")
        self.assertNotIn("eval(", source)


if __name__ == "__main__":
    unittest.main()
