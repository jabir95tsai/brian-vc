# -*- coding: utf-8 -*-
import sys, json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from validate_case_data import validate_data
from prospectus_contract import coverage_excel_rows

SHEET_SKELETON = [
    {"name": "00_封面", "kind": "cover", "default_header": ["項目", "內容"]},
    {"name": "00_覆蓋率", "default_header": ["#", "項目", "狀態", "對應分頁", "來源/說明"]},
    {"name": "01_公司沿革", "default_header": ["年度", "重要紀事", "來源"]},
    {"name": "02_股本形成", "default_header": ["日期", "發行價(元)", "實收股數", "實收金額(元)", "核定股本(股)", "股本來源", "核准文號", "來源"]},
    {"name": "03_董監事", "default_header": ["職稱", "姓名", "性別/年齡", "國籍/註冊地", "初次選任", "本次選任", "任期", "選任時持股(股)", "持股比率%", "主要經(學)歷", "目前兼任其他公司職務", "來源"]},
    {"name": "04_經營團隊", "default_header": ["職稱", "姓名", "性別", "國籍", "就任日期", "持有股數", "持股比率%", "配偶/未成年子女持股(股)", "配偶子女持股%", "主要經(學)歷", "來源"]},
    {"name": "05_主要股東", "default_header": ["主要股東", "持股", "比例%", "備註", "來源"]},
    {"name": "06_產品營收比重", "default_header": ["項目", "前期金額", "前期%", "近期金額", "近期%", "來源"]},
    {"name": "07_目前產品與服務", "default_header": ["類別", "項目", "說明", "來源"]},
    {"name": "08_未來新產品服務", "default_header": ["方向", "新產品/服務", "說明", "來源"]},
    {"name": "09_未來研發計畫", "default_header": ["項目", "內容", "來源"]},
    {"name": "10_研發成功技術", "default_header": ["年份", "開發項目", "來源"]},
    {"name": "11_研發人員學經歷", "default_header": ["研發人員學歷", "前期人數", "前期%", "近期人數", "近期%", "來源"]},
    {"name": "12_從業員工人數", "default_header": ["項目", "前期", "近期", "最新", "來源"]},
    {"name": "13_產品用途", "default_header": ["產品", "主要用途", "來源"]},
    {"name": "14_產製過程", "default_header": ["項目", "內容", "來源/註記"]},
    {"name": "15_毛利率變化", "default_header": ["項目", "前期", "近期", "變動率%", "來源"]},
    {"name": "16_銷售地區", "default_header": ["區域", "前期金額", "前期%", "近期金額", "近期%", "來源"]},
    {"name": "17_主要供應商", "default_header": ["供應商(占進貨≥10%,常代號)", "前期金額", "前期%", "近期金額", "近期%", "與發行人關係", "來源"]},
    {"name": "18_主要客戶", "default_header": ["客戶(占銷貨≥10%,常代號)", "前期金額", "前期%", "近期金額", "近期%", "與發行人關係", "來源"]},
    {"name": "19_銷售量值", "default_header": ["項目", "前期內銷量", "前期內銷值", "前期外銷量", "前期外銷值", "近期內銷量", "近期內銷值", "近期外銷量", "近期外銷值", "來源"]},
    {"name": "20_生產量值", "default_header": ["項目", "前期產能", "前期產量", "前期產值", "近期產能", "近期產量", "近期產值", "來源"]},
    {"name": "21_產能利用率", "default_header": ["項目", "前期產能", "前期產量", "前期利用率%", "近期產能", "近期產量", "近期利用率%", "來源"]},
    {"name": "22_產業概況", "default_header": ["項目", "內容", "來源/註記"]},
    {"name": "23_中下游關聯", "default_header": ["層級", "內容/關鍵廠商", "來源"]},
    {"name": "24_競爭情勢", "default_header": ["項目", "內容/關鍵字", "來源"]},
    {"name": "25_轉投資", "default_header": ["轉投資事業", "主要營業", "投資成本", "帳面價值", "持股%", "投資損益", "來源"]},
    {"name": "26_重要契約", "default_header": ["契約性質", "當事人", "起訖日期", "主要內容", "限制條款", "來源"]},
    {"name": "27_歷年財務摘要5年", "default_header": ["項目", "5年期間各年...", "來源"]},
    {"name": "28甲_損益表", "default_header": ["損益表科目(仟元;EPS元)", "合併近期", "合併前期", "個體近期", "個體前期", "來源"]},
    {"name": "28乙_資產負債表", "default_header": ["資產負債表科目(仟元)", "合併近期", "合併前期", "個體近期", "個體前期", "來源"]},
    {"name": "28丙_現金流量表", "default_header": ["現金流量表科目(仟元)", "合併近期", "合併前期", "個體近期", "個體前期", "來源"]},
    {"name": "29_AI專家", "default_header": ["專家", "分析內容", "來源"]},
    {"name": "查核意見_明細", "default_header": ["年度", "事務所", "會計師", "查核意見", "來源"]},
    {"name": "紅旗", "default_header": ["#", "紅旗", "說明", "來源"]},
]
SKELETON_NAMES = [s["name"] for s in SHEET_SKELETON]

F = "Arial"
HDR = PatternFill("solid", fgColor="1F3864")
SUB = PatternFill("solid", fgColor="D9E1F2")
MISS = PatternFill("solid", fgColor="FCE4D6")
NA = PatternFill("solid", fgColor="E7E6E6")
tn = Side(style="thin", color="BFBFBF")
BD = Border(left=tn, right=tn, top=tn, bottom=tn)

def hf(): return Font(name=F, size=11, bold=True, color="FFFFFF")
def nf(b=False, c="000000"): return Font(name=F, size=10, bold=b, color=c)
def col(i):
    s = ""; i += 1
    while i:
        i, r = divmod(i - 1, 26); s = chr(65 + r) + s
    return s

def render_sheet(ws, sh):
    header = sh.get("header", []); no_header = sh.get("no_header", False); rows = sh.get("rows", [])
    n = len(header) or (len(rows[0]) if rows else 1); start = 1
    if header: ws.append(header)
    for row in rows: ws.append(list(row))
    last = ws.max_row
    if header and not no_header:
        for j in range(1, n + 1):
            c = ws.cell(row=1, column=j); c.fill = HDR; c.font = hf(); c.border = BD
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        start = 2
    top = sh.get("top", False)
    for r in range(start, last + 1):
        for j in range(1, n + 1):
            c = ws.cell(row=r, column=j); c.border = BD
            if c.font is None or c.font.name != F: c.font = nf()
            c.alignment = Alignment(vertical=("top" if top else "center"), wrap_text=True)
    for j in sh.get("numcols", []):
        for r in range(2, last + 1): ws.cell(row=r, column=j).number_format = "#,##0;(#,##0);-"
    for j in sh.get("pct_cols", []):
        for r in range(2, last + 1): ws.cell(row=r, column=j).number_format = "0.00;(0.00)"
    secs = set(sh.get("section_rows", []))
    if secs:
        for r in range(2, last + 1):
            if ws.cell(row=r, column=1).value in secs:
                ws.cell(row=r, column=1).fill = SUB; ws.cell(row=r, column=1).font = nf(b=True)
    for i, w in enumerate(sh.get("widths", [])): ws.column_dimensions[col(i)].width = w
    nr = last + 2
    for note in sh.get("notes", []):
        ws.cell(row=nr, column=1, value=note).font = nf(); nr += 1

def render_unavailable(ws, sk, status, reason):
    header = sk.get("default_header", ["項目", "內容"]); n = len(header); ws.append(header)
    for j in range(1, n + 1):
        c = ws.cell(row=1, column=j); c.fill = HDR; c.font = hf(); c.border = BD
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    if status == "not_applicable":
        marker = "⚪ 不適用" + (f"（{reason}）" if reason else "")
        fill = NA
        font_color = "666666"
    else:
        marker = "🔴 查無資料／須補件" + (f"（{reason}）" if reason else "")
        fill = MISS
        font_color = "C00000"
    ws.append([marker] + [""] * (n - 1))
    for j in range(1, n + 1):
        c = ws.cell(row=2, column=j); c.fill = fill; c.border = BD; c.font = nf(b=True, c=font_color)
        c.alignment = Alignment(vertical="center", wrap_text=True)
    if n > 1: ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n)
    ws.column_dimensions["A"].width = 28
    for i in range(1, n): ws.column_dimensions[col(i)].width = 16

def build(data, out):
    validation = validate_data(data)
    if validation["status"] != "success":
        raise ValueError(
            "case_data validation failed:\n" + "\n".join(validation["errors"])
        )
    wb = Workbook(); wb.remove(wb.active)
    data_sheets = {s["name"]: s for s in data.get("sheets", []) if s.get("name")}
    coverage_rows = coverage_excel_rows(data["coverage"])
    data_sheets["00_覆蓋率"] = {
        **data_sheets["00_覆蓋率"],
        "header": ["#", "項目", "狀態", "對應分頁", "來源/說明"],
        "rows": coverage_rows,
    }
    emitted = []
    for sk in SHEET_SKELETON:
        name = sk["name"]; ws = wb.create_sheet(name[:31]); sh = data_sheets.get(name)
        if sh and sh.get("status") in ("complete", "partial"): render_sheet(ws, sh)
        else:
            status = sh.get("status", "missing") if sh else "missing"
            reason = sh.get("reason") if sh else "case_data 未提供該母版分頁"
            render_unavailable(ws, sk, status, reason)
        emitted.append(name)
    for s in data.get("sheets", []):
        nm = s.get("name")
        if nm and nm not in SKELETON_NAMES and nm not in emitted:
            ws = wb.create_sheet(nm[:31]); render_sheet(ws, s); emitted.append(nm)
    wb.save(out); return wb.sheetnames

if __name__ == "__main__":
    with open(sys.argv[1], encoding="utf-8") as handle:
        data = json.load(handle)
    names = build(data, sys.argv[2])
    missing_skel = [n for n in SKELETON_NAMES if n not in names]
    print("SAVED", len(names), "sheets ->", sys.argv[2])
    print("OK 35 強制母版分頁齊全" if not missing_skel else ("WARNING 缺母版分頁:"+str(missing_skel)))
