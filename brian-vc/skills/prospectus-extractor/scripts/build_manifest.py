# -*- coding: utf-8 -*-
"""Build the self-excluding QA manifest for prospectus-extractor outputs."""

from __future__ import annotations

import argparse
import hashlib
import json
import zipfile
from datetime import datetime, timezone
from pathlib import Path
from typing import Any
from urllib.parse import urlparse
from xml.etree import ElementTree

from prospectus_contract import (
    CASE_DATA_SCHEMA_VERSION,
    CONTRACT_VERSION,
    EXPECTED_BUSINESS_OUTPUT_SUFFIXES,
    SHEET_SKELETON_NAMES,
    coverage_excel_rows,
)
from factbase_from_case import render_text as render_factbase
from pdf_backend import PdfBackendError, count_pdf_pages
from render_case_markdown import render_coverage, render_raw
from validate_case_data import validate_data

EXCEL_ERROR_TOKENS = ("#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#N/A", "#NUM!", "#NULL!")
OOXML_NS = {"m": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def pdf_page_count(path: Path) -> tuple[int | None, str | None]:
    try:
        return count_pdf_pages(path), None
    except PdfBackendError as exc:
        return None, str(exc)


def inspect_xlsx(path: Path, expected_coverage: list[list[str]] | None = None) -> dict[str, Any]:
    result: dict[str, Any] = {
        "status": "failed",
        "sheets": 0,
        "skeleton_complete": False,
        "skeleton_order_ok": False,
        "coverage_consistent": None,
        "error_tokens": [],
        "errors": [],
    }
    try:
        with zipfile.ZipFile(path) as archive:
            workbook_xml = archive.read("xl/workbook.xml")
            root = ElementTree.fromstring(workbook_xml)
            names = [node.attrib["name"] for node in root.findall("m:sheets/m:sheet", OOXML_NS)]
            result["sheets"] = len(names)
            result["sheet_names"] = names
            result["skeleton_complete"] = all(name in names for name in SHEET_SKELETON_NAMES)
            result["skeleton_order_ok"] = names[: len(SHEET_SKELETON_NAMES)] == SHEET_SKELETON_NAMES
            found: list[str] = []
            for member in archive.namelist():
                if member.endswith(".xml"):
                    text = archive.read(member).decode("utf-8", errors="ignore")
                    found.extend(token for token in EXCEL_ERROR_TOKENS if token in text)
            result["error_tokens"] = sorted(set(found))
    except (OSError, KeyError, zipfile.BadZipFile, ElementTree.ParseError) as exc:
        result["errors"].append(str(exc))
        return result

    if expected_coverage is not None:
        try:
            from openpyxl import load_workbook

            workbook = load_workbook(path, read_only=True, data_only=False)
            try:
                worksheet = workbook["00_覆蓋率"]
                actual = [
                    ["" if value is None else value for value in row]
                    for row in worksheet.iter_rows(
                        min_row=2,
                        max_row=1 + len(expected_coverage),
                        min_col=1,
                        max_col=5,
                        values_only=True,
                    )
                ]
                result["coverage_consistent"] = actual == expected_coverage
                if not result["coverage_consistent"]:
                    result["errors"].append(
                        "Excel 00_覆蓋率 does not match case_data.coverage"
                    )
            finally:
                workbook.close()
        except (ImportError, KeyError, OSError, ValueError) as exc:
            result["coverage_consistent"] = False
            result["errors"].append(f"cannot verify Excel coverage: {exc}")

    coverage_ok = result["coverage_consistent"] is not False
    if (
        result["skeleton_complete"]
        and result["skeleton_order_ok"]
        and not result["error_tokens"]
        and coverage_ok
    ):
        result["status"] = "success"
    return result


def source_snapshot_path(source: dict[str, Any], base_dir: Path) -> Path | None:
    candidate = source.get("snapshot_path") or source.get("path_or_url")
    if not candidate:
        return None
    parsed = urlparse(str(candidate))
    if parsed.scheme in ("http", "https"):
        return None
    path = Path(candidate)
    return path if path.is_absolute() else (base_dir / path).resolve()


def build_manifest(case_data_path: Path, output_dir: Path) -> dict[str, Any]:
    data = json.loads(case_data_path.read_text(encoding="utf-8"))
    contract_qa = validate_data(data)
    case_id = data.get("case", {}).get("case_id", case_data_path.stem.replace("_case_data", ""))

    inputs = []
    input_ok = True
    for source in data.get("sources", []):
        snapshot = source_snapshot_path(source, case_data_path.parent)
        entry = {
            "source_id": source.get("source_id"),
            "role": source.get("role"),
            "path_or_url": source.get("path_or_url"),
            "snapshot_path": str(snapshot) if snapshot else None,
            "exists": bool(snapshot and snapshot.is_file()),
            "sha256": None,
            "bytes": None,
            "pages": None,
            "page_count_backend": None,
            "errors": [],
        }
        if snapshot and snapshot.is_file():
            entry["sha256"] = sha256_file(snapshot)
            entry["bytes"] = snapshot.stat().st_size
            if snapshot.suffix.lower() == ".pdf":
                entry["pages"], page_error = pdf_page_count(snapshot)
                entry["page_count_backend"] = "pypdf" if entry["pages"] else None
                if page_error:
                    entry["errors"].append(page_error)
                if source.get("role") in ("prospectus", "prospectus_revision") and not entry["pages"]:
                    input_ok = False
        else:
            input_ok = False
            entry["errors"].append("source snapshot does not exist")
        inputs.append(entry)

    outputs = []
    outputs_exist = True
    outputs_match = contract_qa["status"] == "success"
    excel_qa: dict[str, Any] | None = None
    render_errors: list[str] = []
    expected_bytes: dict[str, bytes] = {}
    if contract_qa["status"] == "success":
        try:
            expected_bytes = {
                "_case_data.json": case_data_path.read_bytes(),
                "_Prospectus_raw.md": render_raw(data).encode("utf-8"),
                "_Factbase.md": render_factbase(data).encode("utf-8"),
                "_Prospectus_coverage.md": render_coverage(data).encode("utf-8"),
            }
        except Exception as exc:
            outputs_match = False
            render_errors.append(f"cannot render expected Markdown: {exc}")
    for suffix in EXPECTED_BUSINESS_OUTPUT_SUFFIXES:
        path = output_dir / f"{case_id}{suffix}"
        exists = path.is_file()
        expected = expected_bytes.get(suffix)
        content_match = (path.read_bytes() == expected) if exists and expected is not None else None
        entry = {
            "path": str(path.resolve()),
            "exists": exists,
            "sha256": sha256_file(path) if exists else None,
            "bytes": path.stat().st_size if exists else None,
            "expected_sha256": hashlib.sha256(expected).hexdigest() if expected is not None else None,
            "content_match": content_match,
        }
        outputs.append(entry)
        outputs_exist = outputs_exist and exists
        if content_match is False:
            outputs_match = False
        if suffix.endswith(".xlsx") and exists:
            excel_qa = inspect_xlsx(path, coverage_excel_rows(data["coverage"]))

    qa = {
        "case_data": contract_qa,
        "inputs_available": input_ok,
        "coverage_exactly_24": len(data.get("coverage", [])) == 24,
        "outputs_complete": outputs_exist,
        "rendered_outputs_match": outputs_match,
        "render_errors": render_errors,
        "excel": excel_qa or {"status": "failed", "errors": ["Excel output missing"]},
    }
    success = (
        contract_qa["status"] == "success"
        and input_ok
        and qa["coverage_exactly_24"]
        and outputs_exist
        and outputs_match
        and qa["excel"]["status"] == "success"
    )
    return {
        "contract_version": CONTRACT_VERSION,
        "case_data_schema_version": CASE_DATA_SCHEMA_VERSION,
        "case_id": case_id,
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "manifest_hash_policy": "self_excluded_to_avoid_recursive_hash",
        "inputs": inputs,
        "outputs": outputs,
        "qa": qa,
        "validation_status": "success" if success else "failed",
    }


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("case_data")
    parser.add_argument("manifest_out")
    parser.add_argument("--output-dir")
    args = parser.parse_args()

    case_data_path = Path(args.case_data).resolve()
    output_dir = Path(args.output_dir).resolve() if args.output_dir else case_data_path.parent
    manifest = build_manifest(case_data_path, output_dir)
    Path(args.manifest_out).write_text(
        json.dumps(manifest, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    print(json.dumps(manifest, ensure_ascii=False, indent=2))
    raise SystemExit(0 if manifest["validation_status"] == "success" else 2)


if __name__ == "__main__":
    main()
