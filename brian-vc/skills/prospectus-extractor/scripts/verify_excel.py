# -*- coding: utf-8 -*-
"""verify_excel.py — 檢查 35 母版、錯誤字串、狀態標記與 cell equality。"""
import sys, json
from openpyxl import load_workbook
ERR = ("#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#N/A", "#NUM!", "#NULL!")
SKELETON_NAMES = [
    "00_封面", "00_覆蓋率", "01_公司沿革", "02_股本形成", "03_董監事", "04_經營團隊",
    "05_主要股東", "06_產品營收比重", "07_目前產品與服務", "08_未來新產品服務",
    "09_未來研發計畫", "10_研發成功技術", "11_研發人員學經歷", "12_從業員工人數",
    "13_產品用途", "14_產製過程", "15_毛利率變化", "16_銷售地區", "17_主要供應商",
    "18_主要客戶", "19_銷售量值", "20_生產量值", "21_產能利用率", "22_產業概況",
    "23_中下游關聯", "24_競爭情勢", "25_轉投資", "26_重要契約", "27_歷年財務摘要5年",
    "28甲_損益表", "28乙_資產負債表", "28丙_現金流量表", "29_AI專家",
    "查核意見_明細", "紅旗",
]
def main():
    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8")
    f = sys.argv[1]
    wb = load_workbook(f, data_only=False)
    names = wb.sheetnames
    missing_skel = [n for n in SKELETON_NAMES if n not in names]
    order_ok = names[:len(SKELETON_NAMES)] == SKELETON_NAMES
    errs = []; flagged_missing = []; flagged_not_applicable = []
    for ws in wb.worksheets:
        marker = ws["A2"].value
        if isinstance(marker, str):
            if marker.startswith("🔴 查無資料／須補件"):
                flagged_missing.append(ws.title)
            elif marker.startswith("⚪ 不適用"):
                flagged_not_applicable.append(ws.title)
        for row in ws.iter_rows():
            for c in row:
                if isinstance(c.value, str):
                    for e in ERR:
                        if e in c.value: errs.append(f"{ws.title}!{c.coordinate}:{e}")
    eq_results = []; args = sys.argv[2:]; i = 0
    while i < len(args):
        if args[i] == "--eq" and i + 1 < len(args):
            spec = args[i + 1]
            try:
                sheet, cell, expected_text = spec.split(":", 2)
                if sheet not in wb.sheetnames:
                    raise ValueError(f"unknown worksheet: {sheet}")
                got = wb[sheet][cell].value
                try:
                    expected = float(expected_text)
                    ok = isinstance(got, (int, float)) and abs(float(got) - expected) < 1e-6
                except ValueError:
                    expected = expected_text
                    ok = str(got) == expected
                eq_results.append(
                    {"spec": spec, "sheet": sheet, "cell": cell, "expected": expected, "got": got, "ok": ok}
                )
            except Exception as ex:
                eq_results.append({"spec": spec, "error": str(ex), "ok": False})
            i += 2
        else: i += 1
    skeleton_ok = (not missing_skel) and order_ok
    eq_ok = all(r.get("ok") for r in eq_results)
    ok = skeleton_ok and (not errs) and eq_ok
    out = {"status": "success" if ok else "failed", "sheets": len(names),
        "skeleton_required": len(SKELETON_NAMES), "skeleton_complete": skeleton_ok,
        "skeleton_missing": missing_skel, "skeleton_order_ok": order_ok,
        "extra_sheets": [n for n in names if n not in SKELETON_NAMES],
        "flagged_missing_pages": flagged_missing,
        "flagged_not_applicable_pages": flagged_not_applicable,
        "error_cells": errs, "eq_checks": eq_results}
    print(json.dumps(out, ensure_ascii=False, indent=2))
    sys.exit(0 if ok else 2)
if __name__ == "__main__": main()
